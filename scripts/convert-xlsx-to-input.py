"""
Converte planilha XLSX (formato Sky Brazil) para CSV de input do pipeline.

Uso:
    python scripts/convert-xlsx-to-input.py docs/20260729_Cancelados_may-26.xlsx

A planilha deve ter as colunas:
    - isu_user_sk (UUID do user)
    - from_date (início do período de extração)
    - to_date (fim do período / data de cancelamento)

Gera um CSV compatível com o pipeline:
    user_id,from_date,to_date
    9e527027-...,2026-01-04,2026-04-30
"""

import sys
from datetime import datetime
from pathlib import Path

import openpyxl


def parse_date(value) -> str:
    """Converte valor de data (datetime ou string) para YYYY-MM-DD."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    # Tentar parsear formatos comuns (DD/MM/YYYY, DD-MM-YYYY)
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def convert(xlsx_path: str, output_path: str = None, max_users: int = None):
    """Converte XLSX para CSV de input."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"Erro: arquivo não encontrado: {xlsx_path}")
        sys.exit(1)

    if output_path is None:
        output_path = xlsx_path.with_suffix(".csv")
    output_path = Path(output_path)

    print(f"Lendo: {xlsx_path}")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    # Encontrar índices das colunas pelo header
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Mapear colunas (aceita variações de nome)
    user_col = None
    from_col = None
    to_col = None

    for i, h in enumerate(headers):
        if h in ("isu_user_sk", "user_id", "userid", "subscriber_id"):
            user_col = i
        elif h in ("from_date", "fromdate", "start_date"):
            from_col = i
        elif h in ("to_date", "todate", "end_date", "cancel_date"):
            to_col = i

    if user_col is None:
        print(f"Erro: coluna de user ID não encontrada. Colunas disponíveis: {headers}")
        sys.exit(1)

    print(f"  Colunas: user_id={headers[user_col]}, from_date={headers[from_col] if from_col else 'N/A'}, to_date={headers[to_col] if to_col else 'N/A'}")

    # Processar linhas
    count = 0
    skipped = 0

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write("user_id,from_date,to_date\n")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if max_users and count >= max_users:
                break

            user_id = str(row[user_col]).strip() if row[user_col] else ""
            if not user_id or user_id == "None":
                skipped += 1
                continue

            from_date = parse_date(row[from_col]) if from_col is not None else ""
            to_date = parse_date(row[to_col]) if to_col is not None else ""

            f.write(f"{user_id},{from_date},{to_date}\n")
            count += 1

    wb.close()

    print(f"  Processados: {count} users")
    if skipped:
        print(f"  Ignorados: {skipped} (sem user_id)")
    print(f"  Output: {output_path}")
    print(f"  Tamanho: {output_path.stat().st_size / 1024:.1f} KB")
    print()
    print("Para executar o pipeline:")
    print(f"  aws s3 cp {output_path} s3://sky-brazil-churn-prediction/input/{output_path.name} --region us-east-1")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/convert-xlsx-to-input.py <arquivo.xlsx> [output.csv] [--max N]")
        print()
        print("Exemplos:")
        print("  python scripts/convert-xlsx-to-input.py docs/20260729_Cancelados.xlsx")
        print("  python scripts/convert-xlsx-to-input.py docs/20260729_Cancelados.xlsx input.csv --max 100")
        sys.exit(1)

    xlsx_file = sys.argv[1]
    output_file = None
    max_n = None

    for i, arg in enumerate(sys.argv[2:], 2):
        if arg == "--max" and i + 1 < len(sys.argv):
            max_n = int(sys.argv[i + 1])
        elif not arg.startswith("--"):
            output_file = arg

    convert(xlsx_file, output_file, max_n)
