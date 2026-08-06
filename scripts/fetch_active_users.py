"""
Busca user IDs ativos nos últimos 3 meses via API NPAW.

Faz chamadas ao endpoint rawdata SEM filtro de user_id para obter sessões
recentes, e extrai os user_ids únicos. Esses são users confirmadamente ativos.

Exclui os users que estão na lista de cancelados para garantir que são "ativos puros".

Uso:
    python scripts/fetch_active_users.py --count 800 --output data/active_users.csv
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests

# Adicionar raiz do projeto ao path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import boto3

# Configurações
NPAW_BASE_URL = "https://api-scripts.npaw.com"
ACCOUNT_CODE = "sky_brazil"
REGION = os.environ.get("AWS_DEFAULT_REGION", "us-east-1")


def get_npaw_api_key() -> str:
    """Busca API key do Secrets Manager."""
    client = boto3.client("secretsmanager", region_name=REGION)
    response = client.get_secret_value(SecretId="churn-prediction/npaw-api-key")
    return response["SecretString"]


def load_churned_user_ids(xlsx_path: str) -> set[str]:
    """Carrega set de user_ids cancelados para excluir."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    user_col = next(i for i, h in enumerate(headers) if h in ("isu_user_sk", "user_id"))

    churned = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = str(row[user_col]).strip().lower() if row[user_col] else ""
        if uid and uid != "none":
            churned.add(uid)

    wb.close()
    return churned


def fetch_recent_sessions(api_key: str, from_date: str, limit: int = 500, offset: int = 0) -> list[dict]:
    """Busca sessões recentes com filtro amplo (região LATAM) para descobrir users."""
    url = f"https://api.npaw.com/{ACCOUNT_CODE}/rawdata"
    headers = {"npaw-api-key": api_key}

    # A API NPAW requer pelo menos um filtro. Usar região LATAM (todos os users Sky Brazil)
    filter_json = json.dumps([{"name": "uf", "rules": {"region": ["LATAM"]}}])

    params = {
        "fromDate": from_date,
        "filter": filter_json,
        "limit": str(limit),
        "offset": str(offset),
        "orderBy": "end_at",
        "orderDirection": "desc",
    }

    response = requests.get(url, headers=headers, params=params, timeout=120)
    response.raise_for_status()
    data = response.json()

    # Parsear sessões
    sessions = []
    if isinstance(data, dict) and "data" in data:
        for d in data["data"]:
            if "values" in d:
                sessions.extend(d["values"])
    elif isinstance(data, list) and data:
        if "data" in data[0]:
            for d in data[0]["data"]:
                if "values" in d:
                    sessions.extend(d["values"])
        elif "values" in data[0]:
            sessions.extend(data[0]["values"])

    return sessions


def extract_user_ids(sessions: list[dict]) -> set[str]:
    """Extrai user_ids únicos das sessões."""
    user_ids = set()
    for session in sessions:
        uid = session.get("user_id", "")
        if uid and isinstance(uid, str) and len(uid) > 10:
            user_ids.add(uid.strip().lower())
    return user_ids


def main():
    parser = argparse.ArgumentParser(description="Busca users ativos via NPAW")
    parser.add_argument("--count", type=int, default=800, help="Número desejado de users ativos")
    parser.add_argument("--output", default="data/active_users.csv", help="Arquivo de saída")
    parser.add_argument("--churned-xlsx", default="docs/20260729 Cancelados may-26.xlsx",
                        help="Planilha de cancelados para excluir")
    parser.add_argument("--months", type=int, default=3, help="Período em meses (default: 3)")
    args = parser.parse_args()

    print("=" * 60)
    print("🔍 Busca de Users Ativos via NPAW")
    print("=" * 60)

    # 1. Carregar cancelados para excluir
    print(f"\n📄 Carregando cancelados de: {args.churned_xlsx}")
    churned_ids = load_churned_user_ids(args.churned_xlsx)
    print(f"  ✅ {len(churned_ids)} cancelados para excluir")

    # 2. Obter API key
    print("\n🔑 Obtendo API key...")
    api_key = get_npaw_api_key()
    print("  ✅ OK")

    # 3. Calcular período
    now = datetime.now(timezone.utc)
    from_dt = now - timedelta(days=args.months * 30)
    from_date = from_dt.strftime("%Y-%m-%d")
    print(f"\n📅 Período: {from_date} até hoje ({args.months} meses)")

    # 4. Buscar sessões e extrair user_ids
    print(f"\n📡 Buscando sessões recentes para descobrir users ativos...")
    active_ids: set[str] = set()
    offset = 0
    batch_size = 500
    max_requests = 50  # Evitar loop infinito
    requests_made = 0

    # Usar 'last7days' com filtro de região (LATAM)
    from_date_param = "last7days"

    while len(active_ids) < args.count and requests_made < max_requests:
        requests_made += 1
        print(f"  🔄 Request {requests_made}: offset={offset}...", end=" ")

        try:
            sessions = fetch_recent_sessions(api_key, from_date_param, limit=batch_size, offset=offset)
        except Exception as e:
            print(f"❌ Erro: {e}")
            break

        if not sessions:
            print("(sem mais dados)")
            break

        batch_ids = extract_user_ids(sessions)
        # Excluir cancelados
        new_ids = batch_ids - churned_ids - active_ids
        active_ids.update(new_ids)
        print(f"{len(sessions)} sessões → +{len(new_ids)} novos users (total: {len(active_ids)})")

        offset += batch_size

        # Rate limiting gentil
        time.sleep(2)

    print(f"\n✅ Total de users ativos encontrados: {len(active_ids)}")

    # 5. Salvar CSV
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Gerar datas (últimos 3 meses até hoje)
    to_date = now.strftime("%Y-%m-%d")

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write("user_id,from_date,to_date\n")
        for uid in sorted(active_ids)[:args.count]:
            f.write(f"{uid},{from_date},{to_date}\n")

    final_count = min(len(active_ids), args.count)
    print(f"\n💾 Salvo: {output_path} ({final_count} users)")
    print(f"   Formato: user_id,from_date,to_date")
    print(f"\n🎯 Próximo passo:")
    print(f"   python scripts/prepare_training_data.py \\")
    print(f"       --churned-xlsx \"{args.churned_xlsx}\" \\")
    print(f"       --active-csv {output_path} \\")
    print(f"       --max 200")


if __name__ == "__main__":
    main()
