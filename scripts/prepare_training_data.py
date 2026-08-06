"""
Script para preparar dados de treinamento do modelo de churn.

Etapas:
1. Lê planilha de cancelados (label=1)
2. Lê lista de ativos (label=0) — se disponível
3. Executa pipeline de extração + feature engineering para ambos
4. Gera CSV consolidado: features + label
5. Upload para S3

Uso:
    # Só cancelados (para gerar features dos cancelados primeiro):
    python scripts/prepare_training_data.py --churned-xlsx "docs/20260729 Cancelados may-26.xlsx" --max 200

    # Com ativos:
    python scripts/prepare_training_data.py \
        --churned-xlsx "docs/20260729 Cancelados may-26.xlsx" \
        --active-csv data/active_users.csv \
        --max 200

    # Usar features já extraídas do S3:
    python scripts/prepare_training_data.py --from-s3 --execution-id <UUID>
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# Adicionar raiz do projeto ao path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import boto3

from src.extractors.ingestion import ingest_user_ids
from src.extractors.npaw_extractor import NPAWExtractor
from src.features.feature_engineer import FeatureEngineer


# Configurações
BUCKET = os.environ.get("BUCKET_NAME", "sky-brazil-churn-prediction")
REGION = os.environ.get("AWS_DEFAULT_REGION", "us-east-1")
NPAW_ACCOUNT_CODE = "sky_brazil"


def get_npaw_api_key() -> str:
    """Busca API key do Secrets Manager."""
    client = boto3.client("secretsmanager", region_name=REGION)
    response = client.get_secret_value(SecretId="churn-prediction/npaw-api-key")
    return response["SecretString"]


def load_churned_users(xlsx_path: str, max_users: int | None = None) -> list[dict]:
    """Carrega users cancelados da planilha XLSX.

    Returns:
        Lista de dicts: {user_id, from_date, to_date, label}
    """
    import openpyxl

    print(f"📄 Lendo planilha: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Encontrar índices
    user_col = next(i for i, h in enumerate(headers) if h in ("isu_user_sk", "user_id"))
    from_col = next((i for i, h in enumerate(headers) if h in ("from_date", "fromdate")), None)
    to_col = next((i for i, h in enumerate(headers) if h in ("to_date", "todate")), None)

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if max_users and len(users) >= max_users:
            break

        user_id = str(row[user_col]).strip() if row[user_col] else ""
        if not user_id or user_id == "None":
            continue

        from_date = ""
        to_date = ""
        if from_col is not None and row[from_col]:
            fd = row[from_col]
            from_date = fd.strftime("%Y-%m-%d") if isinstance(fd, datetime) else str(fd)
        if to_col is not None and row[to_col]:
            td = row[to_col]
            if isinstance(td, datetime):
                to_date = td.strftime("%Y-%m-%d")
            else:
                # Tentar parsear DD/MM/YYYY
                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        to_date = datetime.strptime(str(td).strip(), fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue

        users.append({
            "user_id": user_id,
            "from_date": from_date,
            "to_date": to_date,
            "label": 1,  # Cancelado
        })

    wb.close()
    print(f"  ✅ {len(users)} cancelados carregados")
    return users


def load_active_users(csv_path: str, max_users: int | None = None) -> list[dict]:
    """Carrega users ativos de um CSV.

    CSV esperado: user_id,from_date,to_date
    """
    print(f"📄 Lendo ativos: {csv_path}")
    users = []

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if max_users and len(users) >= max_users:
                break
            users.append({
                "user_id": row["user_id"].strip(),
                "from_date": row.get("from_date", "").strip(),
                "to_date": row.get("to_date", "").strip(),
                "label": 0,  # Ativo
            })

    print(f"  ✅ {len(users)} ativos carregados")
    return users


async def extract_features_for_users(
    users: list[dict],
    api_key: str,
    batch_size: int = 20,
    max_concurrent: int = 10,
) -> list[dict]:
    """Extrai features NPAW para uma lista de users EM PARALELO.

    Usa o NPAWExtractor com concorrência controlada para acelerar
    a extração. Cada user é processado em paralelo até o limite
    de max_concurrent.

    Returns:
        Lista de dicts com features + label, prontos para CSV.
    """
    extractor = NPAWExtractor(
        account_code=NPAW_ACCOUNT_CODE,
        api_key=api_key,
        rate_limit_seconds=0.5,  # Mais agressivo para treinar
        max_concurrent=max_concurrent,
    )
    engineer = FeatureEngineer()

    results = []
    total = len(users)
    processed = 0
    semaphore = asyncio.Semaphore(max_concurrent)

    async def process_user(user_info: dict) -> dict | None:
        nonlocal processed
        user_id = user_info["user_id"]
        from_date = user_info["from_date"]
        to_date = user_info["to_date"]
        label = user_info["label"]

        async with semaphore:
            try:
                sessions = await extractor.extract_user_sessions(
                    user_id=user_id,
                    from_date=from_date,
                    to_date=to_date,
                )

                if not sessions:
                    processed += 1
                    return None

                fv = engineer.compute(user_id=user_id, sessions=sessions)
                if fv is None:
                    processed += 1
                    return None

                fv_dict = fv.model_dump()
                fv_dict["label"] = label
                processed += 1
                return fv_dict

            except Exception as e:
                processed += 1
                return None

    # Processar em batches para mostrar progresso
    for i in range(0, total, batch_size):
        batch = users[i:i + batch_size]
        batch_num = i // batch_size + 1
        total_batches = (total + batch_size - 1) // batch_size
        print(f"  🔄 Batch {batch_num}/{total_batches} ({len(batch)} users em paralelo)...")

        tasks = [process_user(u) for u in batch]
        batch_results = await asyncio.gather(*tasks)

        for r in batch_results:
            if r is not None:
                results.append(r)

        print(f"    ✅ {len(results)} features gerados até agora ({processed}/{total} processados)")

    return results


def features_to_training_csv(features: list[dict]) -> str:
    """Converte lista de features em CSV de treinamento.

    Remove campos não-numéricos e mantém apenas features + label.
    """
    if not features:
        raise ValueError("Nenhuma feature disponível para treinamento")

    # Colunas numéricas do FeatureVector (excluir metadados string)
    exclude_cols = {
        "user_id", "version", "generated_at",
        "observation_start", "observation_end",
    }

    # Usar primeiro registro para determinar colunas
    all_cols = [k for k in features[0].keys() if k not in exclude_cols]

    # Garantir que 'label' está no final
    if "label" in all_cols:
        all_cols.remove("label")
    all_cols.append("label")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()

    for fv in features:
        # Converter None para 0 (trends nulas)
        row = {}
        for col in all_cols:
            val = fv.get(col)
            row[col] = val if val is not None else 0.0
        writer.writerow(row)

    return output.getvalue()


def upload_training_data(csv_content: str, s3_key: str) -> str:
    """Upload CSV de treinamento para S3."""
    s3 = boto3.client("s3", region_name=REGION)
    s3.put_object(
        Bucket=BUCKET,
        Key=s3_key,
        Body=csv_content.encode("utf-8"),
        ContentType="text/csv",
    )
    s3_uri = f"s3://{BUCKET}/{s3_key}"
    print(f"  ✅ Upload: {s3_uri}")
    return s3_uri


def main():
    parser = argparse.ArgumentParser(description="Prepara dados de treinamento para o modelo de churn")
    parser.add_argument("--churned-xlsx", required=True, help="Planilha de cancelados")
    parser.add_argument("--active-csv", default=None, help="CSV de users ativos (opcional)")
    parser.add_argument("--max", type=int, default=None, help="Máximo de users por grupo")
    parser.add_argument("--batch-size", type=int, default=10, help="Tamanho do batch de extração")
    parser.add_argument("--output", default=None, help="Path local para salvar o CSV (opcional)")
    parser.add_argument("--skip-upload", action="store_true", help="Não fazer upload para S3")
    args = parser.parse_args()

    print("=" * 60)
    print("🚀 Preparação de Dados de Treinamento - Churn Prediction")
    print("=" * 60)
    start_time = time.time()

    # 1. Carregar users
    churned_users = load_churned_users(args.churned_xlsx, args.max)

    active_users = []
    if args.active_csv:
        active_users = load_active_users(args.active_csv, args.max)

    all_users = churned_users + active_users
    print(f"\n📊 Total: {len(all_users)} users ({len(churned_users)} cancelados, {len(active_users)} ativos)")

    # 2. Obter API key
    print("\n🔑 Obtendo API key do Secrets Manager...")
    api_key = get_npaw_api_key()
    print("  ✅ API key obtida")

    # 3. Extrair features
    print("\n📡 Extraindo features da NPAW...")
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        features = loop.run_until_complete(
            extract_features_for_users(all_users, api_key, args.batch_size)
        )
    finally:
        loop.close()

    print(f"\n📊 Features extraídos: {len(features)} users com dados")
    churned_with_data = sum(1 for f in features if f["label"] == 1)
    active_with_data = sum(1 for f in features if f["label"] == 0)
    print(f"  - Cancelados com dados: {churned_with_data}")
    print(f"  - Ativos com dados: {active_with_data}")

    if not features:
        print("\n❌ Nenhuma feature extraída. Verifique a API NPAW e os user IDs.")
        sys.exit(1)

    # 4. Gerar CSV
    print("\n📝 Gerando CSV de treinamento...")
    csv_content = features_to_training_csv(features)
    lines = csv_content.count("\n")
    print(f"  ✅ {lines - 1} registros, {len(csv_content) / 1024:.1f} KB")

    # 5. Salvar localmente (opcional)
    if args.output:
        Path(args.output).write_text(csv_content, encoding="utf-8")
        print(f"  💾 Salvo: {args.output}")

    # 6. Upload para S3
    if not args.skip_upload:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        s3_key = f"training_data/churn_training_{timestamp}.csv"
        print(f"\n☁️  Upload para S3...")
        s3_uri = upload_training_data(csv_content, s3_key)
        print(f"\n🎯 Para treinar o modelo, execute:")
        print(f"   python scripts/train_model.py --data-s3 {s3_uri}")

    duration = time.time() - start_time
    print(f"\n✅ Concluído em {duration:.0f}s")


if __name__ == "__main__":
    main()
