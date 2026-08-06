"""
Extração RÁPIDA de features para treinamento.

Diferente do prepare_training_data.py, este script:
- Faz requests diretamente com aiohttp (sem overhead do extractor)
- Limita a 100 sessões por user (suficiente para features)
- Roda 20 requests em paralelo
- Não tem rate limiting artificial

Uso:
    python scripts/fast_training_extract.py --max 200

Tempo estimado: ~3-5 min para 400 users (200 cancelados + 200 ativos)
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import aiohttp
import boto3

from src.features.feature_engineer import FeatureEngineer


# Config
NPAW_URL = "https://api.npaw.com/sky_brazil/rawdata"
REGION = "us-east-1"
BUCKET = "sky-brazil-churn-prediction"
MAX_CONCURRENT = 20  # 20 requests simultâneos
SESSIONS_PER_USER = 100  # Suficiente para calcular features


def get_api_key() -> str:
    client = boto3.client("secretsmanager", region_name=REGION)
    resp = client.get_secret_value(SecretId="churn-prediction/npaw-api-key")
    return resp["SecretString"]


def load_churned(xlsx_path: str, max_n: int) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    user_col = next(i for i, h in enumerate(headers) if h in ("isu_user_sk", "user_id"))
    from_col = next((i for i, h in enumerate(headers) if h == "from_date"), None)
    to_col = next((i for i, h in enumerate(headers) if h == "to_date"), None)

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(users) >= max_n:
            break
        uid = str(row[user_col]).strip() if row[user_col] else ""
        if not uid or uid == "None":
            continue
        from_date = ""
        to_date = ""
        if from_col and row[from_col]:
            fd = row[from_col]
            from_date = fd.strftime("%Y-%m-%d") if isinstance(fd, datetime) else str(fd)
        if to_col and row[to_col]:
            td = row[to_col]
            if isinstance(td, datetime):
                to_date = td.strftime("%Y-%m-%d")
            else:
                for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                    try:
                        to_date = datetime.strptime(str(td).strip(), fmt).strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue
        users.append({"user_id": uid, "from_date": from_date, "to_date": to_date, "label": 1})
    wb.close()
    return users


def load_active(csv_path: str, max_n: int) -> list[dict]:
    users = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if len(users) >= max_n:
                break
            users.append({
                "user_id": row["user_id"].strip(),
                "from_date": row.get("from_date", "").strip(),
                "to_date": row.get("to_date", "").strip(),
                "label": 0,
            })
    return users


async def fetch_user_sessions(
    session: aiohttp.ClientSession,
    user_id: str,
    from_date: str,
    to_date: str,
    api_key: str,
    semaphore: asyncio.Semaphore,
) -> list[dict]:
    """Busca até 100 sessões de um user via NPAW."""
    async with semaphore:
        filter_json = json.dumps([{"name": "uf", "rules": {"user_id": [user_id]}}])
        params = {
            "fromDate": from_date or "last6months",
            "filter": filter_json,
            "limit": str(SESSIONS_PER_USER),
            "offset": "0",
            "orderBy": "end_at",
            "orderDirection": "desc",
        }
        if to_date:
            params["toDate"] = to_date

        headers = {"npaw-api-key": api_key}

        try:
            async with session.get(NPAW_URL, params=params, headers=headers, timeout=aiohttp.ClientTimeout(total=60)) as resp:
                if resp.status != 200:
                    return []
                data = await resp.json()

                # Parsear
                sessions = []
                if isinstance(data, dict) and "data" in data:
                    for d in data["data"]:
                        if "values" in d:
                            sessions.extend(d["values"])
                return sessions
        except Exception:
            return []


async def extract_all(users: list[dict], api_key: str) -> list[dict]:
    """Extrai features de todos os users em paralelo."""
    engineer = FeatureEngineer()
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    results = []
    total = len(users)
    done = 0

    connector = aiohttp.TCPConnector(limit=MAX_CONCURRENT, ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        # Processar em batches de 50 para mostrar progresso
        batch_size = 50
        for i in range(0, total, batch_size):
            batch = users[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (total + batch_size - 1) // batch_size

            tasks = []
            for u in batch:
                tasks.append(fetch_user_sessions(
                    session, u["user_id"], u["from_date"], u["to_date"], api_key, semaphore
                ))

            sessions_list = await asyncio.gather(*tasks)

            for u, sessions in zip(batch, sessions_list):
                done += 1
                if not sessions:
                    continue
                try:
                    fv = engineer.compute(user_id=u["user_id"], sessions=sessions)
                    if fv:
                        fv_dict = fv.model_dump()
                        fv_dict["label"] = u["label"]
                        results.append(fv_dict)
                except Exception:
                    continue

            print(f"  Batch {batch_num}/{total_batches}: {done}/{total} processados, {len(results)} com features")

    return results


def to_csv(features: list[dict]) -> str:
    """Converte features para CSV de treinamento."""
    exclude = {"user_id", "version", "generated_at", "observation_start", "observation_end"}
    all_cols = [k for k in features[0].keys() if k not in exclude]
    if "label" in all_cols:
        all_cols.remove("label")
    all_cols.append("label")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=all_cols, extrasaction="ignore")
    writer.writeheader()
    for fv in features:
        row = {col: (fv.get(col) if fv.get(col) is not None else 0.0) for col in all_cols}
        writer.writerow(row)
    return output.getvalue()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--max", type=int, default=200)
    parser.add_argument("--churned-xlsx", default="docs/20260729 Cancelados may-26.xlsx")
    parser.add_argument("--active-csv", default="data/active_users.csv")
    parser.add_argument("--output", default="data/training_data_sample.csv")
    args = parser.parse_args()

    print("=" * 60)
    print("⚡ Extração RÁPIDA de Features para Treinamento")
    print(f"   Paralelismo: {MAX_CONCURRENT} requests simultâneos")
    print(f"   Sessões/user: {SESSIONS_PER_USER}")
    print("=" * 60)
    start = time.time()

    # Carregar users
    print(f"\n📄 Carregando cancelados...")
    churned = load_churned(args.churned_xlsx, args.max)
    print(f"  {len(churned)} cancelados")

    print(f"📄 Carregando ativos...")
    active = load_active(args.active_csv, args.max)
    print(f"  {len(active)} ativos")

    all_users = churned + active
    print(f"\n📊 Total: {len(all_users)} users ({len(churned)} churn + {len(active)} ativo)")

    # API key
    print("\n🔑 Obtendo API key...")
    api_key = get_api_key()

    # Extrair
    print(f"\n📡 Extraindo features ({MAX_CONCURRENT} em paralelo)...")
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        features = loop.run_until_complete(extract_all(all_users, api_key))
    finally:
        loop.close()

    churn_count = sum(1 for f in features if f["label"] == 1)
    active_count = sum(1 for f in features if f["label"] == 0)
    print(f"\n✅ Features extraídos: {len(features)} ({churn_count} churn, {active_count} ativo)")

    if not features:
        print("❌ Nenhuma feature extraída!")
        sys.exit(1)

    # Gerar CSV
    csv_content = to_csv(features)
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.output).write_text(csv_content, encoding="utf-8")
    print(f"\n💾 CSV salvo: {args.output} ({len(features)} registros)")

    # Upload S3
    print("\n☁️  Upload para S3...")
    s3 = boto3.client("s3", region_name=REGION)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    s3_key = f"training_data/churn_training_{ts}.csv"
    s3.put_object(Bucket=BUCKET, Key=s3_key, Body=csv_content.encode(), ContentType="text/csv")
    print(f"  ✅ s3://{BUCKET}/{s3_key}")

    elapsed = time.time() - start
    print(f"\n🏁 Concluído em {elapsed:.0f}s ({elapsed/60:.1f} min)")
    print(f"\n🎯 Próximo passo - treinar modelo:")
    print(f"   python scripts/train_model.py --data-s3 s3://{BUCKET}/{s3_key}")


if __name__ == "__main__":
    main()
